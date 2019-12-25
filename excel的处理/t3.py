# -*- coding：UTF-8 -*-
from openpyxl import load_workbook
from pprint import pprint


class ExcelOperate(object):
    '''
        excel的操作相关
    '''

    def get_data(self, file_path, sheet_name):
        '''
        读取excel的数据，并返回
        :param file_path:
        :param sheet_name:
        :return:data[list]
        '''
        wb = load_workbook(file_path)
        a_sheet = wb[sheet_name]
        max_row = a_sheet.max_row
        max_column = a_sheet.max_column
        # 循环读取数据（dict）,并放入data（list）中 ，row 是行
        data = []
        for i in range(2, max_row + 1):
            data_dict = {}
            for j in range(1, max_column + 1):
                title = a_sheet.cell(row=1, column=j).value
                value = a_sheet.cell(i, j).value
                data_dict[title] = value
            data.append(data_dict)
        return data

    def edit_excel(self, file_path, sheet_name, change_list):
        '''
        将编号在change_list中的行的第change_column列
        的值修改为flag
        :param file_path: 文件目录
        :param sheet_name: sheet页名
        :param change_list: 修改列表
        :return:
        '''
        wb = load_workbook(file_path)
        a_sheet = wb[sheet_name]
        # max_row = a_sheet.max_row
        # max_column = a_sheet.max_column
        rows = a_sheet.rows
        # column = a_sheet.columns
        change_column = 7

        # a = a_sheet['A2'].value
        for row in rows:

            # 每一行
            line = [col.value for col in row]
            print(line)
            # print(line[0])

            if line[0] in change_list:
                a_sheet.cell(row=int(line[0]), column=change_column).value = 'flag'
                wb.save(file_path)
        print('更新列表：{}'.format(change_list))


if __name__ == '__main__':
    file_path = 'test1.xlsx'
    sheet_name = 's1'
    change_list = {5, 6}
    eo = ExcelOperate()
    data = eo.get_data(file_path, sheet_name)
    pprint(data)
    # data = eo.edit_excel(file_path, sheet_name, change_list)
    # pprint(data)
